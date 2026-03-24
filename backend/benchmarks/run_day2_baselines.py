"""
Day 2 baseline benchmark runner.
Runs no-correction + random baselines and writes results.xlsx

Usage:
    cd backend
    python benchmarks/run_day2_baselines.py --dataset_id 1
"""

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from core.database import SessionLocal
from services.benchmark_service import (
    run_no_correction_benchmark,
    run_random_benchmark,
    get_benchmark_results,
)


def run(dataset_id: int):
    print(f"\n{'='*60}")
    print(f"  DAY 2 BASELINES — Dataset {dataset_id}")
    print(f"{'='*60}\n")

    db = SessionLocal()
    try:
        # 1. No correction
        print("Running no-correction baseline...")
        no_corr = run_no_correction_benchmark(db, dataset_id)
        print(f"  Accuracy: {no_corr['accuracy']:.4f}\n")

        # 2. Random correction
        print("Running random correction baseline...")
        rand = run_random_benchmark(db, dataset_id)
        print(f"  Accuracy: {rand['accuracy']:.4f}")
        print(f"  Samples flagged: {rand['human_effort']}\n")

        # 3. Write results.xlsx
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            print("openpyxl not installed. Run: pip install openpyxl")
            sys.exit(1)

        output_path = Path(__file__).parent / "results.xlsx"

        # Load existing or create new
        if output_path.exists():
            wb = openpyxl.load_workbook(output_path)
        else:
            wb = openpyxl.Workbook()
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        # Create/overwrite Dataset 1 sheet
        sheet_name = f"Dataset_{dataset_id}"
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)

        # Headers
        headers = ["Tool", "Iteration", "Accuracy", "Precision", "Recall", "F1", "Human Effort"]
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Fetch all results from DB
        all_results = get_benchmark_results(db, dataset_id)

        # Write rows
        tool_fills = {
            "no_correction": "FCE4D6",
            "random": "FFF2CC",
            "cleanlab": "DDEBF7",
            "sldce": "E2EFDA",
        }

        row = 2
        for r in all_results:
            fill_color = tool_fills.get(r["tool"], "FFFFFF")
            row_fill = PatternFill("solid", fgColor=fill_color)

            values = [
                r["tool"],
                r["iteration"],
                round(r["accuracy"] * 100, 2) if r["accuracy"] else None,
                round(r["precision"] * 100, 2) if r["precision"] else None,
                round(r["recall"] * 100, 2) if r["recall"] else None,
                round(r["f1"] * 100, 2) if r["f1"] else None,
                r["human_effort"],
            ]

            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = row_fill
                if col in [3, 4, 5, 6] and val is not None:
                    cell.number_format = "0.00%"

            row += 1

        # Auto-width columns
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 12)

        wb.save(output_path)
        print(f"Results saved to: {output_path}")

        # Print summary table
        print(f"\n{'='*60}")
        print("  SUMMARY (all tools run so far)")
        print(f"{'='*60}")
        print(f"  {'Tool':<20} {'Iter':<6} {'Accuracy':<12} {'F1':<10}")
        print(f"  {'-'*50}")
        for r in all_results:
            acc = f"{r['accuracy']*100:.2f}%" if r['accuracy'] else "—"
            f1 = f"{r['f1']*100:.2f}%" if r['f1'] else "—"
            print(f"  {r['tool']:<20} {r['iteration']:<6} {acc:<12} {f1:<10}")
        print(f"{'='*60}\n")

    finally:
        db.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Run Day 2 baselines")
    parser.add_argument("--dataset_id", type=int, required=True)
    args = parser.parse_args()
    run(args.dataset_id)